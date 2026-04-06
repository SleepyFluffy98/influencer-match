"""
Streamlit app — Brand intake form (Tab 1), Results dashboard (Tab 2),
and My Preferences analytics (Tab 3).
Run with: streamlit run frontend/app.py
Mock mode: set USE_MOCK=true in .env to run without real API keys.
"""
import io
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv()

import streamlit as st

# On Streamlit Community Cloud there is no .env file.
# Inject st.secrets into os.environ so all service modules pick them up via os.getenv().
try:
    for _k, _v in st.secrets.items():
        if isinstance(_v, str):
            os.environ.setdefault(_k, _v)
except Exception:
    pass  # Local dev — st.secrets may not exist; .env already loaded above

# ── Password gate ─────────────────────────────────────────────────────────────
def _check_password() -> bool:
    """Return True if the user has entered the correct password."""
    correct = os.getenv("APP_PASSWORD", "")
    if not correct:
        return True  # No password configured — open access (local dev)

    if st.session_state.get("authenticated"):
        return True

    st.title("Influencer Match")
    pwd = st.text_input("Password", type="password", key="pwd_input")
    if st.button("Enter"):
        if pwd == correct:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not _check_password():
    st.stop()
# ─────────────────────────────────────────────────────────────────────────────

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db.store import list_jobs, load_job, new_job, save_job, update_status
from db.feedback import (
    REJECTION_REASONS,
    archive_feedback_log,
    get_feedback_stats,
    save_feedback,
)
from services.preference_learner import build_preference_context
from services.scorer import ScoredInfluencer
from services.hashtag_generator import BrandBrief
from services.fetchers.youtube import COUNTRY_CODES

_USE_MOCK = os.getenv("USE_MOCK", "false").lower() == "true"

if _USE_MOCK:
    from services.mocks import fetch_profiles, score_profiles
else:
    from services.fetcher import fetch_profiles
    from services.scorer import score_profiles

# ── Constants ─────────────────────────────────────────────────────────────────

TABLE_COLS   = [4, 1.5, 1.5, 2.2, 1.8, 3.5]

TIER_LABELS_IG = {
    "nano":  "Nano (5k – 50k)",
    "micro": "Micro (50k – 100k)",
    "macro": "Macro (100k+)",
}
TIER_LABELS_YT = {
    "nano":  "Nano (1k – 10k)",
    "micro": "Micro (10k – 100k)",
    "macro": "Macro (100k+)",
}

# ── Pure helpers ──────────────────────────────────────────────────────────────

def _initials(username: str) -> str:
    words = username.replace(".", " ").replace("_", " ").split()
    return "".join(w[0].upper() for w in words if w)[:2]


def _fmt_followers(n: int) -> str:
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1_000:
        return f"{n / 1_000:.1f}k"
    return str(n)


def _niche_tag(profile) -> str:
    for word in (profile.bio or "").split():
        if word.startswith("#") and len(word) > 2:
            return word[1:12]
    return profile.username.replace(".", "").replace("_", "")[:10]


def _fmt_date(iso: str | None) -> str:
    """Format an ISO datetime string to a short date, or return '—'."""
    if not iso:
        return "—"
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%-d %b %Y")
    except ValueError:
        return "—"


def _status_color(status: str) -> tuple[str, str]:
    if status == "approved":
        return "#EAF3DE", "#27500A"
    if status == "rejected":
        return "#FCEBEB", "#791F1F"
    if status == "maybe":
        return "#FFF3CD", "#856404"
    return "#FAEEDA", "#633806"


def _to_xlsx(scored: list[ScoredInfluencer]) -> bytes:
    """
    Serialize all influencer results to a formatted XLSX file.
    Includes every data field. Returns raw bytes for st.download_button.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Influencer Results"

    headers = [
        "Username", "Platform", "Full Name", "Followers", "Following",
        "Posts / Videos", "Engagement Rate", "Avg Views (YT)", "Last Posted",
        "Bio", "Profile URL",
        "Audience Match", "Niche Relevance", "Engagement Quality", "Brand Safety",
        "Overall Score", "Status", "Rationale", "Rejection Reason", "Notes",
    ]

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    alt_fill = PatternFill("solid", fgColor="F5F5F5")

    for row_idx, s in enumerate(scored, start=2):
        p = s.profile
        row_fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            p.username,
            p.platform,
            p.full_name,
            p.followers,
            p.following,
            p.posts_count,
            f"{p.engagement_rate:.1%}",
            p.avg_views if p.avg_views is not None else "",
            _fmt_date(p.last_posted_at),
            p.bio,
            p.profile_url,
            s.audience_match,
            s.niche_relevance,
            s.engagement_quality,
            s.brand_safety,
            s.overall_score,
            s.status,
            s.rationale,
            s.rejection_reason or "",
            s.notes or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Make profile URL a clickable hyperlink
        url_col = headers.index("Profile URL") + 1
        url_cell = ws.cell(row=row_idx, column=url_col)
        if p.profile_url:
            url_cell.hyperlink = p.profile_url
            url_cell.font = Font(color="4A90D9", underline="single")
            url_cell.value = p.profile_url

    # Auto-fit column widths (cap at 60 characters)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



def _save_note(job_id: str, username: str, note_text: str) -> None:
    """Update just the notes field for an influencer in the job JSON (no JSONL append)."""
    job = load_job(job_id)
    if not job:
        return
    updated = [
        s.model_copy(update={"notes": note_text or None}) if s.profile.username == username else s
        for s in job.results
    ]
    update_status(job_id, job.status, results=updated)


def _inject_css() -> None:
    st.markdown("""
    <style>
    .th {
        font-size: 11px; font-weight: 700;
        text-transform: uppercase; color: #999;
        letter-spacing: 0.06em; padding: 4px 0;
    }
    .cell { font-size: 14px; color: #333; padding: 10px 0; }
    .avatar {
        width: 36px; height: 36px; border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        color: white; font-weight: 700; font-size: 13px; flex-shrink: 0;
    }
    .badge {
        display: inline-block; padding: 3px 10px;
        border-radius: 20px; font-size: 12px; font-weight: 600;
    }
    .rationale-box {
        background: #F0F7FF; border-left: 4px solid #4A90D9;
        padding: 12px 16px; border-radius: 4px; margin-top: 20px;
        display: flex; gap: 10px; align-items: flex-start;
    }
    .rationale-label { font-weight: 700; font-size: 13px; color: #1A5276; white-space: nowrap; }
    .rationale-text  { font-size: 13px; color: #2C3E50; }
    </style>
    """, unsafe_allow_html=True)


def _render_actions(s: ScoredInfluencer, job_id: str) -> None:
    """
    Render action controls for one row.
    • Pending: Approve / Maybe / Reject buttons.
    • Reject: inline reason picker + optional note + Confirm/Cancel.
    • Decided: show result + "↩ Change" to re-decide + optional note for approved/maybe.
    """
    username = s.profile.username
    status   = s.status
    key      = f"{username}__{job_id}"

    reject_key   = f"rejecting_{key}"
    noting_key   = f"noting_{key}"
    changing_key = f"changing_{key}"

    # "Change" mode — show action buttons even if already decided
    in_change_mode = st.session_state.get(changing_key, False)

    if status == "pending" or in_change_mode:
        if st.session_state.get(reject_key):
            # ── Reject form ───────────────────────────────────────────
            reason = st.selectbox(
                "Reason", REJECTION_REASONS,
                key=f"reason_{key}", label_visibility="collapsed",
            )
            note = st.text_input(
                "note", placeholder="Add note...",
                key=f"note_input_{key}", label_visibility="collapsed",
            )
            c_confirm, c_cancel = st.columns(2)
            with c_confirm:
                if st.button("Confirm reject", key=f"confirm_{key}", type="primary"):
                    if reason == REJECTION_REASONS[0]:
                        st.warning("Pick a reason first.")
                    else:
                        save_feedback(job_id, username, "rejected", reason, note or None)
                        st.session_state.pop(reject_key, None)
                        st.session_state.pop(changing_key, None)
                        st.rerun()
            with c_cancel:
                if st.button("Cancel", key=f"cancel_{key}"):
                    st.session_state.pop(reject_key, None)
                    st.session_state.pop(changing_key, None)
                    st.rerun()
        else:
            # ── Action buttons ────────────────────────────────────────
            b1, b2, b3 = st.columns(3)
            with b1:
                if st.button("✓ Approve", key=f"app_{key}", use_container_width=True):
                    save_feedback(job_id, username, "approved", None, None)
                    st.session_state.pop(changing_key, None)
                    st.session_state[noting_key] = True
                    st.rerun()
            with b2:
                if st.button("? Maybe", key=f"maybe_{key}", use_container_width=True):
                    save_feedback(job_id, username, "maybe", None, None)
                    st.session_state.pop(changing_key, None)
                    st.session_state[noting_key] = True
                    st.rerun()
            with b3:
                if st.button("✗ Reject", key=f"rej_{key}", use_container_width=True):
                    st.session_state[reject_key] = True
                    st.rerun()

    elif status == "rejected":
        reason_text = s.rejection_reason or "—"
        st.markdown(
            f'<div style="font-size:11px;color:#791F1F;padding:2px 0">'
            f'<b>Reason:</b> {reason_text}</div>',
            unsafe_allow_html=True,
        )
        if st.button("↩ Change", key=f"change_{key}"):
            st.session_state[changing_key] = True
            st.rerun()

    else:
        # Approved / Maybe — show optional notes
        if st.session_state.get(noting_key):
            saved_note = st.text_input(
                "note", value=s.notes or "",
                placeholder="Add a note (optional)...",
                key=f"note_field_{key}", label_visibility="collapsed",
            )
            if st.button("Save note", key=f"save_note_{key}"):
                _save_note(job_id, username, saved_note)
                st.session_state.pop(noting_key, None)
                st.rerun()
        else:
            if s.notes:
                st.markdown(
                    f'<div style="font-size:11px;color:#555;padding:2px 0">📝 {s.notes}</div>',
                    unsafe_allow_html=True,
                )
            c_note, c_change = st.columns(2)
            with c_note:
                if st.button("+ Note", key=f"open_note_{key}", use_container_width=True):
                    st.session_state[noting_key] = True
                    st.rerun()
            with c_change:
                if st.button("↩ Change", key=f"change_{key}", use_container_width=True):
                    st.session_state[changing_key] = True
                    st.rerun()


def _render_row(s: ScoredInfluencer, job_id: str) -> None:
    """Render one influencer row: styled info cells + action controls."""
    username = s.profile.username
    score    = s.overall_score
    status   = s.status

    opacity = "0.55" if score < 65 else "1"
    status_bg, status_fg = _status_color(status)

    if status == "approved":
        avatar_bg = "#D6EDCA"
    elif status == "rejected":
        avatar_bg = "#FADBD8"
    elif status == "maybe":
        avatar_bg = "#FFF3CD"
    else:
        avatar_bg = "#E0E0E0"

    cols = st.columns(TABLE_COLS)

    if s.profile.platform == "youtube":
        pill_bg, pill_fg, pill_label = "#FFE5E5", "#CC0000", "YT"
    else:
        pill_bg, pill_fg, pill_label = "#FFE5F0", "#CC0066", "IG"

    # ── Col 0: Influencer identity ───────────────────────────────────────────
    last_posted_html = ""
    if s.profile.last_posted_at:
        last_posted_html = (
            f'<div style="font-size:10px;color:#aaa;margin-top:1px">'
            f'Last posted {_fmt_date(s.profile.last_posted_at)}</div>'
        )

    with cols[0]:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:10px;opacity:{opacity};padding:8px 0">
            <div class="avatar" style="background:{avatar_bg}">{_initials(username)}</div>
            <div>
                <div style="display:flex;align-items:center;gap:6px">
                    <span style="background:{pill_bg};color:{pill_fg};font-size:10px;
                                 font-weight:700;padding:2px 6px;border-radius:4px">{pill_label}</span>
                    <a href="{s.profile.profile_url}" target="_blank" rel="noopener noreferrer"
                       style="font-weight:600;font-size:14px;color:#222;text-decoration:none">
                        @{username} ↗</a>
                </div>
                <div style="font-size:11px;color:#888;margin-top:1px">#{_niche_tag(s.profile)}</div>
                {last_posted_html}
            </div>
        </div>""", unsafe_allow_html=True)

    # ── Remaining columns ────────────────────────────────────────────────────
    with cols[1]:
        st.markdown(
            f'<div class="cell" style="opacity:{opacity}">{_fmt_followers(s.profile.followers)}</div>',
            unsafe_allow_html=True,
        )

    with cols[2]:
        st.markdown(
            f'<div class="cell" style="opacity:{opacity}">{s.profile.engagement_rate:.1%}</div>',
            unsafe_allow_html=True,
        )

    with cols[3]:
        st.markdown(
            f'<div class="cell" style="opacity:{opacity}">{score}</div>',
            unsafe_allow_html=True,
        )

    with cols[4]:
        st.markdown(
            f'<div style="opacity:{opacity};padding:10px 0">'
            f'<span class="badge" style="background:{status_bg};color:{status_fg};'
            f'text-transform:capitalize">{status}</span></div>',
            unsafe_allow_html=True,
        )

    with cols[5]:
        _render_actions(s, job_id)


# ── Page ──────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Influencer Matcher", layout="wide")
st.title("Influencer Matcher")

tab_search, tab_results, tab_prefs = st.tabs(["New Search", "Results", "My Preferences"])

# ── Tab 1: New search ─────────────────────────────────────────────────────────

with tab_search:
    st.subheader("Brand Brief")

    # Platform + tier config (outside form so tier dropdowns react to platform selection)
    platforms = st.multiselect(
        "Platforms", ["instagram", "youtube"], default=["instagram"],
        key="platforms_select",
    )

    platform_tiers: dict[str, str] = {}
    if platforms:
        tier_cols = st.columns(len(platforms))
        for col, plat in zip(tier_cols, platforms):
            with col:
                if plat == "instagram":
                    choice = st.selectbox(
                        "Instagram follower tier",
                        list(TIER_LABELS_IG.keys()),
                        format_func=lambda k: TIER_LABELS_IG[k],
                        index=1,
                        key="ig_tier",
                    )
                    platform_tiers["instagram"] = choice
                elif plat == "youtube":
                    choice = st.selectbox(
                        "YouTube follower tier",
                        list(TIER_LABELS_YT.keys()),
                        format_func=lambda k: TIER_LABELS_YT[k],
                        index=1,
                        key="yt_tier",
                    )
                    platform_tiers["youtube"] = choice

    with st.form("brief_form"):
        brand_name    = st.text_input("Brand name", placeholder="e.g. GlowLab")
        industry      = st.selectbox("Industry / niche",
            ["beauty", "fashion", "fitness", "food", "lifestyle", "tech", "travel", "wellness", "other"])
        target_age    = st.selectbox("Target audience age", ["13-17", "18-24", "25-34", "35-44", "45+"])
        target_gender = st.selectbox("Target audience gender", ["female", "male", "all"])
        campaign_goal = st.selectbox("Campaign goal", ["awareness", "conversion", "content"])
        countries     = st.multiselect("Countries / regions (optional)", list(COUNTRY_CODES.keys()),
            placeholder="Global (no filter)")
        keywords      = st.text_area("Keywords / brand values",
            placeholder="e.g. sustainable, clean beauty, minimalist")
        red_flags     = st.text_area("Red flags to avoid",
            placeholder="e.g. no alcohol, no competitor brands")
        contact_email = st.text_input("Contact email")
        submitted     = st.form_submit_button("Run search", type="primary")

    if submitted:
        if not brand_name or not contact_email:
            st.error("Brand name and contact email are required.")
        elif not platforms:
            st.error("Select at least one platform.")
        else:
            default_tier = platform_tiers.get(platforms[0], "micro")
            brief = BrandBrief(
                brand_name=brand_name, industry=industry, target_age=target_age,
                target_gender=target_gender, campaign_goal=campaign_goal,
                follower_tier=default_tier, keywords=keywords,
                red_flags=red_flags, contact_email=contact_email,
                platforms=platforms, countries=countries,
                platform_tiers=platform_tiers,
            )
            job = new_job(brief)
            save_job(job)
            try:
                with st.spinner("Fetching profiles..."):
                    profiles = fetch_profiles(brief)

                if not profiles:
                    update_status(job.job_id, "failed", error="No profiles returned")
                    st.warning("No profiles found — try broader keywords or a different platform.")
                    st.stop()
                else:
                    with st.spinner(f"Scoring {len(profiles)} influencers..."):
                        scored = score_profiles(profiles, brief)
                    update_status(
                        job.job_id, "complete", results=scored,
                        completed_at=datetime.now(timezone.utc).isoformat(),
                    )
                    st.success(f"Done! Found {len(scored)} matches. Switch to the Results tab.")
            except Exception as exc:
                update_status(job.job_id, "failed", error=str(exc))
                st.error(f"Search failed: {exc}")
                st.stop()

# ── Tab 2: Results ────────────────────────────────────────────────────────────

with tab_results:
    jobs = list_jobs()

    if not jobs:
        st.info("No searches yet. Run a search in the New Search tab.")
    else:
        _inject_css()

        job_labels = {
            j.job_id: f"{j.brand_brief.brand_name} — {j.created_at[:10]} ({j.status})"
            for j in jobs
        }
        selected_id = st.selectbox(
            "job", options=list(job_labels.keys()),
            format_func=lambda jid: job_labels[jid],
            label_visibility="collapsed",
        )
        job = next((j for j in jobs if j.job_id == selected_id), None)

        if not job or not job.results:
            st.warning(
                f"Job status: {job.status if job else '?'}. "
                f"{job.error if job and job.error else 'No results yet.'}"
            )
        else:
            # Page header
            h_left, h_right = st.columns([8, 2])
            with h_left:
                st.markdown(
                    f'<h3 style="margin:0 0 4px">{job.brand_brief.brand_name}</h3>'
                    f'<span style="color:#888;font-size:13px">{job.created_at[:10]} · '
                    f'{job.brand_brief.industry} · '
                    f'{", ".join(getattr(job.brand_brief, "countries", [])) or "Global"}</span>',
                    unsafe_allow_html=True,
                )
            with h_right:
                xlsx_bytes = _to_xlsx(job.results)
                st.download_button(
                    f"Export XLSX ({len(job.results)} rows)",
                    data=xlsx_bytes,
                    file_name=f"{job.brand_brief.brand_name}_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

            # Metric cards
            total     = len(job.results)
            matched   = len([s for s in job.results if s.overall_score >= 60])
            avg_score = int(sum(s.overall_score for s in job.results) / total) if total else 0
            pending   = len([s for s in job.results if s.status == "pending"])
            match_rate = f"{int(matched / total * 100)}%" if total else "0%"

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Profiles scored", total)
            m2.metric("Matched (≥60)", matched, delta=match_rate)
            m3.metric("Avg score", avg_score)
            m4.metric("Pending review", pending)

            st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

            # Filter + sort row
            f_col, _, s_col = st.columns([5, 3, 2])
            with f_col:
                status_filter = st.radio(
                    "filter", ["All", "Pending", "Approved", "Maybe", "Rejected"],
                    horizontal=True, label_visibility="collapsed",
                )
            with s_col:
                sort_by = st.selectbox(
                    "sort", ["Score", "Followers", "Engagement"],
                    label_visibility="collapsed",
                )

            # Apply filter + sort
            results = job.results
            if status_filter != "All":
                results = [r for r in results if r.status == status_filter.lower()]
            if sort_by == "Score":
                results = sorted(results, key=lambda s: s.overall_score, reverse=True)
            elif sort_by == "Followers":
                results = sorted(results, key=lambda s: s.profile.followers, reverse=True)
            elif sort_by == "Engagement":
                results = sorted(results, key=lambda s: s.profile.engagement_rate, reverse=True)

            # Table header
            st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
            for col, label in zip(
                st.columns(TABLE_COLS),
                ["Influencer", "Followers", "Engagement", "Score", "Status", "Actions"],
            ):
                with col:
                    st.markdown(f'<div class="th">{label}</div>', unsafe_allow_html=True)
            st.markdown(
                "<hr style='margin:4px 0 0;border:none;border-top:2px solid #E8E8E8'>",
                unsafe_allow_html=True,
            )

            # Table rows
            for s in results:
                _render_row(s, job.job_id)

            # Rationale callout for top result
            if results:
                top = results[0]
                st.markdown(
                    f'<div class="rationale-box">'
                    f'<span class="rationale-label">Top match · @{top.profile.username}</span>'
                    f'<span class="rationale-text">{top.rationale}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

# ── Tab 3: My Preferences ─────────────────────────────────────────────────────

with tab_prefs:
    st.subheader("My Preferences")
    st.caption("Decisions you make train the system to score future results closer to your taste.")

    stats = get_feedback_stats()
    total_decisions = stats["total_decisions"]

    if total_decisions == 0:
        st.info("No decisions recorded yet. Approve or reject influencers in the Results tab.")
    else:
        p1, p2, p3, p4 = st.columns(4)
        p1.metric("Total decisions", total_decisions)
        p2.metric("Approval rate", f"{stats['approval_rate']:.0%}")
        p3.metric("Avg score · approved", stats["avg_score_approved"])
        p4.metric("Avg score · rejected", stats["avg_score_rejected"])

        st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)

        top_reasons = stats["top_rejection_reasons"][:3]
        if top_reasons:
            st.markdown("**Top rejection reasons**")
            st.bar_chart({r: n for r, n in top_reasons})
        else:
            st.markdown("*No rejections recorded yet.*")

        st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)

        st.markdown("**What the system has learned**")
        MIN_DECISIONS = 15
        if total_decisions < MIN_DECISIONS:
            st.info(
                f"Not enough data yet — make {MIN_DECISIONS - total_decisions} more "
                f"decision(s) to activate learning (need {MIN_DECISIONS} total)."
            )
        else:
            with st.spinner("Generating preference summary..."):
                context = build_preference_context(min_decisions=MIN_DECISIONS)
            if context:
                st.markdown(
                    f'<div style="background:#F8F4FF;border-left:4px solid #7C3AED;'
                    f'padding:12px 16px;border-radius:4px;font-size:14px;color:#1A0045">'
                    f'{context}</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.warning("Could not generate preference summary — check Azure OpenAI configuration.")

        st.markdown("<div style='margin-top:32px'></div>", unsafe_allow_html=True)

        st.markdown("**Reset preferences**")
        st.caption("Archives the current feedback log and starts fresh. Past job results are not affected.")
        if st.button("Reset preferences", type="secondary"):
            archive_path = archive_feedback_log()
            if archive_path:
                st.success(f"Preferences reset. Log archived to {archive_path.name}.")
            else:
                st.info("Nothing to reset — feedback log is already empty.")
            st.rerun()
